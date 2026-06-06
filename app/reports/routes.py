import io
from flask import render_template, redirect, url_for, flash, make_response, request
from flask_login import login_required, current_user
from functools import wraps
from datetime import date, timedelta
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont
from app import db
from app.reports import reports
from app.models import Menu, AttendanceConfirmation, User, MenuItem

# ── Colour palette ────────────────────────────────────────────────────────────
_HDR_BG   = '1F4E79'   # dark blue  — header fill
_HDR_FG   = 'FFFFFF'   # white      — header text
_ALT_BG   = 'DCE6F1'   # light blue — alternating row fill
_BORDER   = Side(style='thin', color='B8CCE4')

_HDR_FILL   = PatternFill('solid', fgColor=_HDR_BG)
_ALT_FILL   = PatternFill('solid', fgColor=_ALT_BG)
_HDR_FONT   = Font(name='Calibri', bold=True,  color=_HDR_FG, size=11)
_BODY_FONT  = Font(name='Calibri', bold=False, color='000000', size=10)
_HDR_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
_BODY_ALIGN = Alignment(horizontal='left',   vertical='center')
_CELL_BORDER = Border(
    left=_BORDER, right=_BORDER, top=_BORDER, bottom=_BORDER
)


def _make_watermark_png():
    """Return a BytesIO PNG: tiled diagonal 'UEAB CAFETERIA' in light gray."""
    W, H = 1400, 1000

    # Load the best available bold font; fall back gracefully
    font = None
    for path in [
        'C:/Windows/Fonts/ariblk.ttf',
        'C:/Windows/Fonts/arialbd.ttf',
        'C:/Windows/Fonts/arial.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    ]:
        try:
            font = ImageFont.truetype(path, 80)
            break
        except OSError:
            continue
    if font is None:
        font = ImageFont.load_default()

    # Draw text on a scratch layer then rotate it
    scratch = Image.new('RGBA', (1100, 150), (255, 255, 255, 0))
    d = ImageDraw.Draw(scratch)
    d.text((10, 20), 'UEAB CAFETERIA', font=font, fill=(140, 140, 140, 140))
    rotated = scratch.rotate(35, expand=True)
    rw, rh = rotated.size

    # Tile across the canvas
    canvas = Image.new('RGBA', (W, H), (255, 255, 255, 0))
    for y in range(-rh, H + rh, rh + 80):
        for x in range(-rw // 2, W + rw, rw + 20):
            canvas.paste(rotated, (x, y), rotated)

    buf = io.BytesIO()
    canvas.save(buf, format='PNG')
    buf.seek(0)
    return buf


def _xlsx_response(sheet_title, headers, rows, filename):
    """Build a styled .xlsx workbook and return it as a Flask response."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.freeze_panes = 'A2'   # keep header visible while scrolling

    # ── Watermark ─────────────────────────────────────────────────────────────
    wm = XLImage(_make_watermark_png())
    wm.width, wm.height = 1050, 750
    ws.add_image(wm, 'A1')

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    for col_idx, heading in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=heading)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _CELL_BORDER

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 18
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = _BODY_FONT
            cell.alignment = _BODY_ALIGN
            cell.border    = _CELL_BORDER
            if fill:
                cell.fill = fill

    # ── Auto-fit column widths ────────────────────────────────────────────────
    for col_idx, heading in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(heading))
        for row in rows:
            cell_val = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ''
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Serialize and return ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def staff_or_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if current_user.role not in ('staff', 'admin'):
            flash('Access denied.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if current_user.role != 'admin':
            flash('Access denied.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


@reports.route('/')
@login_required
@staff_or_admin
def index():
    today = date.today()
    week_ago = today - timedelta(days=6)

    daily_rows = (
        db.session.query(Menu.date, Menu.meal_type, func.count(AttendanceConfirmation.id))
        .outerjoin(AttendanceConfirmation, Menu.id == AttendanceConfirmation.menu_id)
        .filter(Menu.date >= week_ago, Menu.date <= today)
        .group_by(Menu.date, Menu.meal_type)
        .order_by(Menu.date)
        .all()
    )

    dates = [(week_ago + timedelta(days=i)).strftime('%b %d') for i in range(7)]
    chart_data = {'breakfast': [0]*7, 'lunch': [0]*7, 'dinner': [0]*7}
    for row_date, meal_type, count in daily_rows:
        delta = (row_date - week_ago).days
        if 0 <= delta < 7 and meal_type in chart_data:
            chart_data[meal_type][delta] = count

    meal_totals = dict(
        db.session.query(Menu.meal_type, func.count(AttendanceConfirmation.id))
        .outerjoin(AttendanceConfirmation, Menu.id == AttendanceConfirmation.menu_id)
        .group_by(Menu.meal_type)
        .all()
    )

    total_students = User.query.filter_by(role='student', is_active=True).count()
    total_confirmations = AttendanceConfirmation.query.count()

    top_attendees = (
        db.session.query(User, func.count(AttendanceConfirmation.id).label('total'))
        .join(AttendanceConfirmation, User.id == AttendanceConfirmation.user_id)
        .filter(User.role == 'student')
        .group_by(User.id)
        .order_by(func.count(AttendanceConfirmation.id).desc())
        .limit(10)
        .all()
    )

    return render_template('reports/index.html',
        dates=dates,
        chart_data=chart_data,
        meal_totals=meal_totals,
        total_students=total_students,
        total_confirmations=total_confirmations,
        top_attendees=top_attendees,
        today=today,
    )


@reports.route('/download/attendance')
@login_required
@staff_or_admin
def download_attendance():
    rows = (
        db.session.query(
            Menu.date, Menu.meal_type,
            User.name, User.student_id, User.email,
            AttendanceConfirmation.confirmed_at
        )
        .join(Menu, AttendanceConfirmation.menu_id == Menu.id)
        .join(User, AttendanceConfirmation.user_id == User.id)
        .order_by(Menu.date.desc(), Menu.meal_type, User.name)
        .all()
    )

    headers = ['Date', 'Meal Type', 'Student Name', 'Student ID', 'Email', 'Confirmed At']
    data = [
        [str(r.date), r.meal_type.capitalize(), r.name,
         r.student_id or '', r.email,
         r.confirmed_at.strftime('%Y-%m-%d %H:%M')]
        for r in rows
    ]
    return _xlsx_response(
        'Attendance Report', headers, data,
        f'attendance_report_{date.today()}.xlsx'
    )


@reports.route('/download/today')
@login_required
@staff_or_admin
def download_today():
    today = date.today()
    rows = (
        db.session.query(
            Menu.meal_type, User.name, User.student_id, User.email,
            AttendanceConfirmation.confirmed_at
        )
        .join(Menu, AttendanceConfirmation.menu_id == Menu.id)
        .join(User, AttendanceConfirmation.user_id == User.id)
        .filter(Menu.date == today)
        .order_by(Menu.meal_type, User.name)
        .all()
    )

    headers = ['Meal Type', 'Student Name', 'Student ID', 'Email', 'Confirmed At']
    data = [
        [r.meal_type.capitalize(), r.name, r.student_id or '',
         r.email, r.confirmed_at.strftime('%H:%M')]
        for r in rows
    ]
    return _xlsx_response(
        "Today's Attendance", headers, data,
        f'attendance_today_{today}.xlsx'
    )


@reports.route('/download/menus')
@login_required
@staff_or_admin
def download_menus():
    menus = Menu.query.order_by(Menu.date.desc(), Menu.meal_type).all()

    headers = ['Date', 'Meal Type', 'Description', 'Menu Items', 'Confirmations']
    data = [
        [
            str(m.date),
            m.meal_type.capitalize(),
            m.description or '',
            ', '.join(i.name for i in m.items),
            m.confirmed_count(),
        ]
        for m in menus
    ]
    return _xlsx_response(
        'Menu Schedule', headers, data,
        f'menu_schedule_{date.today()}.xlsx'
    )


@reports.route('/download/users')
@login_required
@admin_required
def download_users():
    users = User.query.order_by(User.role, User.name).all()

    headers = ['Name', 'Email', 'Student ID', 'Role', 'Status', 'Date Joined']
    data = [
        [u.name, u.email, u.student_id or '', u.role.capitalize(),
         'Active' if u.is_active else 'Inactive',
         u.created_at.strftime('%Y-%m-%d')]
        for u in users
    ]
    return _xlsx_response(
        'User List', headers, data,
        f'user_list_{date.today()}.xlsx'
    )
